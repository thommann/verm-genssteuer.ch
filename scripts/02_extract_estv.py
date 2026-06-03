#!/usr/bin/env python3
"""
02 — ESTV Vermoegenssteuerstatistik -> die Kern-Datensaetze der Seite.

Liest die 11 ESTV-Jahres-Workbooks (data/raw/estv/, Steuerjahre 2012-2022,
Blatt "CH" = gesamtschweizerische Klassentabelle) und erzeugt:

  src/data/estv_distribution.json   Anzahl + Reinvermoegen je Vermoegensklasse/Jahr
                                    (Kategorien "alle" 2012-2022 und
                                     "unbeschraenkt steuerpflichtig" 2020-2022)
  src/data/estv_kennzahlen.json     daraus berechnete Kennzahlen (Mittel, Median,
                                    Perzentile, Anteile, Gini-Bausteine)
  src/data/calculator_params.json   Jahresparameter des Rechners (f, alpha, x_max …),
                                    Tarif-Defaults (Politik-Vorschlag) + Referenz-Aufkommen
  src/data/calculator_bins.json     170 geom. Bins (5 Mio.-50 Mrd.), Besetzung 2020-2022
  src/data/projektion_cohorts.json  30 Kohorten (5 Mio.-30 Mrd.) fuer die Dynamik 2022

Die Klassengrenzen sind fix (ESTV-Schema). Zwei Dateiformate:
  - 2012-2019: Reinvermoegen in Mio. CHF, nur Kategorie "alle"
  - 2020-2022: Reinvermoegen in CHF, zusaetzlich "unbeschraenkt"/"beschraenkt"

Alle Rechenverfahren sind in docs/METHODIK.md dokumentiert und werden von
scripts/00_reproduce_statistics.py unabhaengig nachgeprueft.
"""
import json
import math
import os
import sys

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW = os.path.join(ROOT, "data", "raw", "estv")
DATA = os.path.join(ROOT, "src", "data")

YEARS = list(range(2012, 2023))
UNB_YEARS = [2020, 2021, 2022]

# Fixes ESTV-Klassenschema (11 Klassen). lo/hi/width in CHF; offene Top-Klasse: hi/width = None.
CLASSES = [
    {"label": "",                          "lo": 0,        "hi": 0,        "width": 0},
    {"label": "> 0 - 50'000",              "lo": 0,        "hi": 50000,    "width": 50000},
    {"label": "> 50'000 - 100'000",        "lo": 50000,    "hi": 100000,   "width": 50000},
    {"label": "> 100'000 - 200'000",       "lo": 100000,   "hi": 200000,   "width": 100000},
    {"label": "> 200'000 - 500'000",       "lo": 200000,   "hi": 500000,   "width": 300000},
    {"label": "> 500'000 - 1'000'000",     "lo": 500000,   "hi": 1000000,  "width": 500000},
    {"label": "> 1'000'000 - 2'000'000",   "lo": 1000000,  "hi": 2000000,  "width": 1000000},
    {"label": "> 2'000'000 - 3'000'000",   "lo": 2000000,  "hi": 3000000,  "width": 1000000},
    {"label": "> 3'000'000 - 5'000'000",   "lo": 3000000,  "hi": 5000000,  "width": 2000000},
    {"label": "> 5'000'000 - 10'000'000",  "lo": 5000000,  "hi": 10000000, "width": 5000000},
    {"label": "> 10'000'000",              "lo": 10000000, "hi": None,     "width": None},
]

RAWFILE = {
    2012: "estv-vermoegen-2012.xlsx", 2013: "estv-vermoegen-2013.xlsx",
    2014: "estv-vermoegen-2014.xlsm", 2015: "estv-vermoegen-2015.xlsm",
    2016: "estv-vermoegen-2016.xlsx", 2017: "estv-vermoegen-2017.xlsx",
    2018: "estv-vermoegen-2018.xlsx", 2019: "estv-vermoegen-2019.xlsx",
    2020: "estv-vermoegen-2020.xlsx", 2021: "estv-vermoegen-2021.xlsx",
    2022: "estv-vermoegen-2022.xlsx",
}


def fail(msg):
    print(f"  [FAIL] {msg}", file=sys.stderr)
    sys.exit(1)


def num(v):
    return v if isinstance(v, (int, float)) else None


# ---------------------------------------------------------------------------
# ESTV-Klassentabelle aus dem Blatt "CH" lesen
# ---------------------------------------------------------------------------
def read_year(year):
    """Liest die 11 Klassenwerte (Anzahl + Reinvermoegen in CHF) eines Steuerjahres.

    Rueckgabe: dict mit 'alle' = (counts[11], wealth[11]) und – ab 2020 –
    'unb' = (counts[11], wealth[11]). Reinvermoegen stets in CHF.
    """
    path = os.path.join(RAW, RAWFILE[year])
    if not os.path.exists(path):
        fail(f"ESTV-Datei fehlt: {path} — zuerst `bash scripts/fetch_sources.sh`.")
    ws = openpyxl.load_workbook(path, data_only=True)["CH"]
    rows = [[ws.cell(r, c).value for c in range(1, 15)] for r in range(1, ws.max_row + 1)]

    if year <= 2019:
        # Altes Format: Klassen-Block beginnt bei der Zeile, deren Spalte C == 0
        # ist (Klasse "0"). Anzahl = Spalte D (4), Reinvermoegen = Spalte F (6),
        # in Mio. CHF -> *1e6.
        start = next((i for i, row in enumerate(rows)
                      if str(row[2]).strip() == "0" and num(row[3]) and row[3] > 1000), None)
        if start is None:
            fail(f"{year}: Klassen-Block (Spalte C==0) im Blatt CH nicht gefunden.")
        block = rows[start:start + 11]
        counts = [num(r[3]) for r in block]
        wealth = [num(r[5]) * 1e6 if num(r[5]) is not None else None for r in block]
        return {"alle": (counts, wealth)}

    # Neues Format (2020-2022): Klassen-Block beginnt bei Spalte B (2) == "0".
    # Anzahl alle = C(3), unbeschraenkt = E(5); Reinvermoegen alle = I(9),
    # unbeschraenkt = K(11) – bereits in CHF.
    start = next((i for i, row in enumerate(rows)
                  if str(row[1]).strip() == "0" and num(row[2])), None)
    if start is None:
        fail(f"{year}: Klassen-Block (Spalte B==0) im Blatt CH nicht gefunden.")
    block = rows[start:start + 11]
    return {
        "alle": ([num(r[2]) for r in block], [num(r[8]) for r in block]),
        "unb":  ([num(r[4]) for r in block], [num(r[10]) for r in block]),
    }


def build_distribution():
    raw = {y: read_year(y) for y in YEARS}

    # Selbstpruefung: Klassensumme ~ "Totale"-Zeile bzw. interne Konsistenz.
    for y in YEARS:
        c, w = raw[y]["alle"]
        if any(v is None for v in c):
            fail(f"{y}: fehlende Anzahl-Werte {c}")
        if y in UNB_YEARS:
            cu, wu = raw[y]["unb"]
            # alle = unbeschraenkt + beschraenkt -> unbeschraenkt < alle in jeder Klasse
            if not all(cu[i] <= c[i] + 1 for i in range(11)):
                fail(f"{y}: unbeschraenkt > alle in einer Klasse")

    def col(cat, kind):  # kind: 0=counts, 1=wealth
        years = UNB_YEARS if cat == "unb" else YEARS
        out = []
        for k in range(11):
            out.append({
                "label": CLASSES[k]["label"], "lo": CLASSES[k]["lo"],
                "hi": CLASSES[k]["hi"], "width": CLASSES[k]["width"],
                "values": {str(y): raw[y][cat][kind][k] for y in years},
            })
        return out

    return {
        "alle_counts": col("alle", 0), "alle_wealth": col("alle", 1),
        "unb_counts": col("unb", 0),   "unb_wealth": col("unb", 1),
    }


# ---------------------------------------------------------------------------
# Verteilungs-Kennzahlen aus klassierten Daten (Verfahren A/B/C, docs/METHODIK.md)
# ---------------------------------------------------------------------------
def kennzahlen(counts_rows, wealth_rows, years):
    out = {}
    lo = [c["lo"] for c in counts_rows]
    hi = [c["hi"] for c in counts_rows]
    width = [c["width"] for c in counts_rows]
    for yr in years:
        cnt = [c["values"][str(yr)] for c in counts_rows]
        wl = [c["values"][str(yr)] for c in wealth_rows]
        N, W = sum(cnt), sum(wl)

        def percentile(p):
            target, cum = p * N, 0
            for i in range(len(cnt)):
                if cum + cnt[i] >= target:
                    if not width[i]:
                        return lo[i]
                    return lo[i] + (target - cum) / cnt[i] * width[i]
                cum += cnt[i]
            return lo[-1]

        out[str(yr)] = {
            "N": N, "W": W, "mean": W / N,
            "median": percentile(0.5), "p90": percentile(0.9),
            "p95": percentile(0.95), "p99": percentile(0.99),
            "share_ge10M": wl[10] / W,
            "share_ge5M": (wl[9] + wl[10]) / W,
            "share_ge1M": sum(wl[6:]) / W,
            "cnt_ge10M": cnt[10], "cnt_ge5M": cnt[9] + cnt[10], "cnt_ge1M": sum(cnt[6:]),
            "pct_ge10M": cnt[10] / N, "pct_ge5M": (cnt[9] + cnt[10]) / N,
            "pct_ge1M": sum(cnt[6:]) / N,
        }
    return out


# ---------------------------------------------------------------------------
# Rechner: Jahresparameter, Populationsmodell (Verfahren B/D), Tarif (Verfahren E)
# ---------------------------------------------------------------------------
XMIN = 1e7  # Untergrenze der offenen Top-Klasse


def year_params(dist, m_pauschal):
    """f, alpha, N_tail, x_max je Jahr – allein aus ESTV (unbeschraenkt) + FDK-M."""
    years = {}
    for yr in UNB_YEARS:
        f5_10 = dist["unb_counts"][9]["values"][str(yr)]
        f10 = dist["unb_counts"][10]["values"][str(yr)]
        w10 = dist["unb_wealth"][10]["values"][str(yr)]
        mean10 = w10 / f10
        alpha = mean10 / (mean10 - XMIN)          # Pareto aus Klassenmittel
        ntail = f10 + m_pauschal                  # + Pauschalbesteuerte (FDK)
        xmax = XMIN * ntail ** (1.0 / alpha)      # erwartetes Maximum von N Ziehungen
        years[str(yr)] = {
            "f5_10": f5_10, "f10": f10, "w10": w10, "mean10": mean10,
            "alpha": alpha, "Ntail": ntail, "xmax": xmax,
        }
    return years


def population(params_year, edges):
    """Personen je Bin: 5-10 Mio. gleichverteilt + >10 Mio. Pareto, bei x_max gekappt."""
    f, N, alpha, xmax = (params_year["f5_10"], params_year["Ntail"],
                         params_year["alpha"], params_year["xmax"])
    out = []
    for i in range(len(edges) - 1):
        a, b = edges[i], edges[i + 1]
        cnt = 0.0
        lo5, hi5 = max(a, 5e6), min(b, 1e7)
        if hi5 > lo5:
            cnt += f * (hi5 - lo5) / 5e6
        loT, hiT = max(a, XMIN), min(b, xmax)
        if hiT > loT:
            cnt += N * ((XMIN / loT) ** alpha - (XMIN / hiT) ** alpha)
        out.append(cnt)
    return out


def make_tax(d):
    k = d["exponent"]
    kp = k + 1
    basis = d["ankerSatz"] * d["ankerVermoegen"] * kp / (
        d["schwelle"] * ((d["ankerVermoegen"] / d["schwelle"]) ** kp - 1))
    wcap = d["schwelle"] * (d["cap"] / basis) ** (1.0 / k)

    def tax(W):
        if W <= d["schwelle"]:
            return 0.0
        if W <= wcap:
            return basis * d["schwelle"] / kp * ((W / d["schwelle"]) ** kp - 1)
        below = basis * d["schwelle"] / kp * ((wcap / d["schwelle"]) ** kp - 1)
        return below + d["cap"] * (W - wcap)

    return tax


def build_calculator(dist, m_pauschal):
    defaults = {
        "schwelle": 5000000, "exponent": 0.9, "cap": 1,
        "ankerVermoegen": 100000000, "ankerSatz": 0.02,
        "mPauschal": m_pauschal, "rendite": 0.05,
    }
    years = year_params(dist, m_pauschal)

    # 170 geometrische Bins 5 Mio. .. 50 Mrd.
    NB, LO, HI = 170, 5e6, 5e10
    ratio = (HI / LO) ** (1 / NB)
    edges = [LO * ratio ** i for i in range(NB + 1)]
    pops = {yr: population(years[str(yr)], edges) for yr in UNB_YEARS}
    bins = [{
        "lo": edges[i], "hi": edges[i + 1], "mid": math.sqrt(edges[i] * edges[i + 1]),
        "cnt2020": pops[2020][i], "cnt2021": pops[2021][i], "cnt2022": pops[2022][i],
    } for i in range(NB)]

    # Referenz-Aufkommen des Modells bei Default-Parametern (vom Skript berechnet,
    # keine externe Publikation – dient als Regressions-Anker fuer 00_reproduce).
    tax = make_tax(defaults)
    published = {str(yr): sum(b[f"cnt{yr}"] * tax(b["mid"]) for b in bins) / 1e9
                 for yr in UNB_YEARS}

    # 30 Kohorten 5 Mio. .. 30 Mrd. fuer die dynamische Projektion (Jahr 2022).
    NC, CLO, CHI = 30, 5e6, 3e10
    cratio = (CHI / CLO) ** (1 / NC)
    cedges = [CLO * cratio ** i for i in range(NC + 1)]
    cpop = population(years["2022"], cedges)
    cohorts = [{
        "von": cedges[i], "bis": cedges[i + 1],
        "W0": math.sqrt(cedges[i] * cedges[i + 1]), "anzahl": cpop[i],
    } for i in range(NC)]

    params = {"defaults": defaults, "years": years, "published_revenue": published}
    return params, bins, cohorts


def dump(name, obj, indent):
    with open(os.path.join(DATA, name), "w") as fh:
        json.dump(obj, fh, indent=indent)


def main():
    pausch = json.load(open(os.path.join(DATA, "pauschal.json")))
    m_pauschal = pausch["counts_ch"]["2018"]  # FDK 2018 -> M im Tail

    dist = build_distribution()
    dump("estv_distribution.json", dist, 1)

    kenn = {
        "unbeschraenkt": kennzahlen(dist["unb_counts"], dist["unb_wealth"], UNB_YEARS),
        "alle": kennzahlen(dist["alle_counts"], dist["alle_wealth"], YEARS),
    }
    dump("estv_kennzahlen.json", kenn, 1)

    params, bins, cohorts = build_calculator(dist, m_pauschal)
    dump("calculator_params.json", params, 2)
    dump("calculator_bins.json", bins, 0)
    dump("projektion_cohorts.json", cohorts, 1)

    k22 = kenn["unbeschraenkt"]["2022"]
    print("  [OK ] estv_distribution.json, estv_kennzahlen.json, "
          "calculator_params.json, calculator_bins.json, projektion_cohorts.json")
    print(f"        2022 unbeschraenkt: N={k22['N']:,} median=CHF {k22['median']:,.0f} "
          f"Anteil>=10Mio={k22['share_ge10M']:.4f}  (M_pauschal={m_pauschal})")
    for yr in UNB_YEARS:
        print(f"        Referenz-Aufkommen {yr}: {params['published_revenue'][str(yr)]:.4f} Mrd. "
              f"(alpha={params['years'][str(yr)]['alpha']:.5f})")


if __name__ == "__main__":
    main()
