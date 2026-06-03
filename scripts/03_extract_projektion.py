import openpyxl, json, math
base='/root/.claude/uploads/131eac8a-39f6-47e5-a401-bed7eef8fb5e/'
OUT='/home/user/verm-genssteuer.ch/data/'
rech=openpyxl.load_workbook(base+'1a6139ec-Vermoegenssteuer_Rechner.xlsx', data_only=True)
pj=rech['Projektion']
cohorts=[]
for r in range(6,36):
    if pj.cell(r,1).value is None: continue
    cohorts.append({'von':pj.cell(r,2).value,'bis':pj.cell(r,3).value,'W0':pj.cell(r,4).value,'anzahl':pj.cell(r,5).value})
json.dump(cohorts, open(OUT+'projektion_cohorts.json','w'), indent=1)
print("cohorts:",len(cohorts))
# published summary
pub={}
for j,yr in enumerate(range(2022,2033)):
    pub[yr]=pj.cell(39,2+j).value
print("published dyn revenue:", {k:round(v,2) for k,v in pub.items()})

# Replicate dynamic engine with default params
sr=rech['Steuerrechner']
schwelle=sr['C4'].value;k=sr['C5'].value;cap=sr['C6'].value;anker=sr['C7'].value;target=sr['C8'].value;r_=sr['C10'].value
kp=k+1; basis=target*anker*kp/(schwelle*((anker/schwelle)**kp-1)); wcap=schwelle*(cap/basis)**(1.0/k)
def tax(W):
    if W<=schwelle: return 0.0
    if W<=wcap: return basis*schwelle/kp*((W/schwelle)**kp-1)
    return basis*schwelle/kp*((wcap/schwelle)**kp-1)+cap*(W-wcap)
W=[c['W0'] for c in cohorts]; n=[c['anzahl'] for c in cohorts]
for yr in range(2022,2033):
    rev=sum(n[i]*tax(W[i]) for i in range(len(W)))
    if yr in (2022,2027,2032): print("dyn %d = %.3f bn"%(yr,rev/1e9))
    W=[max(0,W[i]*(1+r_)-tax(W[i])) for i in range(len(W))]
