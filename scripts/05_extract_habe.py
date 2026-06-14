#!/usr/bin/env python3
"""
05: BFS-Haushaltsbudgeterhebung (HABE) nach Einkommensklasse -> src/data/habe.json

Liefert die zwei Haushalts-Bezugsgroessen fuer den Abschnitt «Zucman-Steuer»
(Karte «Und ein normaler Haushalt?»):

  - "arbeiter"     Median-Haushalt = mittleres Einkommensfuenftel (3. Quintil).
  - "mittelstand"  Durchschnitt aller Haushalte (Spalte «Saemtliche»).

Quelle (live, ein einziger Befehl ueber scripts/fetch_sources.sh):
  data/raw/bfs/habe-einkommensklasse.xlsx
  = BFS «Haushaltseinkommen und -ausgaben nach Einkommensklasse», Tabelle
  T20.02.01.00.12, https://dam-api.bfs.admin.ch/hub/api/dam/assets/10867300/master
  Blatt «2015-2017» (neueste Querschnittsperiode). Die fuenf Einkommensklassen
  sind die Quintile der Bruttoeinkommensverteilung (Note [8] der Tabelle), je
  20 % der Haushalte.

Aus dem Blatt werden exakt drei Zeilen je Spalte gelesen (Betrag in CHF/Monat,
Mittelwert):
  - «Bruttoeinkommen»
  - «Einkommen aus Vermoegen und Vermietung»
  - «Steuern» (Einkommens- und Vermoegenssteuern; in der Tabelle negativ -> Betrag)

Daraus folgen die zwei Kennzahlen der Karte (das Bruttoeinkommen kuerzt sich):
  jahre = Steuern / Vermoegenseinkommen        (Zeit, die ganze Jahressteuer
          allein aus dem passiven Vermoegenseinkommen zu verdienen)
  tage  = Steuern / Bruttoeinkommen * 365       (dieselbe Steuersumme aus dem
          gesamten Einkommen, vor allem Arbeit)

Verwendung:
  bash scripts/fetch_sources.sh            # laedt u. a. die HABE-Tabelle
  python3 scripts/05_extract_habe.py
"""
import json
import os
import sys

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "data", "raw", "bfs", "habe-einkommensklasse.xlsx")
OUT = os.path.join(ROOT, "src", "data", "habe.json")
SHEET = "2015–2017"  # neueste Querschnittsperiode

# Zeilen-Beschriftungen exakt wie in der BFS-Tabelle (Spalte 1 bzw. 2).
ROW_BRUTTO = "Bruttoeinkommen"
ROW_VERMOEGEN = "Einkommen aus Vermögen und Vermietung"
ROW_STEUERN = "Steuern"


def fail(msg):
    print(f"  [FAIL] {msg}", file=sys.stderr)
    sys.exit(1)


def find_columns(ws):
    """Header-Zeile mit «Sämtliche» finden und Spalten den Werten zuordnen.

    Die beschrifteten Header-Zellen stehen je eine Spalte rechts von der
    zugehoerigen Wertspalte (z. B. «Sämtliche» in Spalte 5, Werte in Spalte 4).
    Rueckgabe: (saemtliche_col, [quintil_cols links->rechts]).
    """
    for r in range(1, 16):
        labels = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip():
                labels[c] = v.strip()
        if any(t == "Sämtliche" for t in labels.values()):
            saemtliche = None
            quintile = []  # (col, label)
            for c, t in labels.items():
                valcol = c - 1
                if t == "Sämtliche":
                    saemtliche = valcol
                elif any(ch.isdigit() for ch in t):  # Einkommensbaender, z. B. «4 914 – 7 264»
                    quintile.append((valcol, t))
            quintile.sort()
            if saemtliche is None or len(quintile) != 5:
                fail(f"Header unerwartet: Sämtliche={saemtliche}, Quintile={quintile}")
            return saemtliche, [col for col, _ in quintile]
    fail("Header-Zeile mit «Sämtliche» nicht gefunden.")


def find_row(ws, label):
    """Erste Zeile, deren Beschriftung (Spalte 1 oder 2) exakt passt."""
    for r in range(1, ws.max_row + 1):
        for c in (1, 2):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip() == label:
                return r
    fail(f"Zeile «{label}» nicht gefunden.")


def main():
    if not os.path.exists(SRC):
        fail(f"HABE-Datei fehlt: {SRC}; zuerst `bash scripts/fetch_sources.sh`.")
    wb = openpyxl.load_workbook(SRC, data_only=True)
    if SHEET not in wb.sheetnames:
        fail(f"Blatt «{SHEET}» fehlt; vorhanden: {wb.sheetnames}")
    ws = wb[SHEET]

    saemtliche, quintile = find_columns(ws)
    median_col = quintile[2]  # mittleres (3.) Quintil = enthaelt den Median
    top_col = quintile[4]     # oberstes Quintil
    r_brutto = find_row(ws, ROW_BRUTTO)
    r_verm = find_row(ws, ROW_VERMOEGEN)
    r_steuern = find_row(ws, ROW_STEUERN)

    def cell(row, col):
        v = ws.cell(row, col).value
        if not isinstance(v, (int, float)):
            fail(f"Zelle ({row},{col}) ist keine Zahl: {v!r}")
        return float(v)

    def household(col):
        brutto = cell(r_brutto, col)
        vermoegen = abs(cell(r_verm, col))
        steuern = abs(cell(r_steuern, col))
        return {
            "brutto": round(brutto, 1),
            "vermoegen": round(vermoegen, 1),
            "steuern": round(steuern, 1),
            "vermoegen_pct": round(vermoegen / brutto * 100, 2),
            "steuern_pct": round(steuern / brutto * 100, 2),
            # Abgeleitete Kennzahlen der Karte (das Bruttoeinkommen kuerzt sich):
            "jahre": round(steuern / vermoegen, 2),
            "tage": round(steuern / brutto * 365, 1),
        }

    arbeiter = household(median_col)      # Median = mittleres Fuenftel
    mittelstand = household(saemtliche)   # Durchschnitt aller Haushalte
    top_vermoegen_pct = round(abs(cell(r_verm, top_col)) / cell(r_brutto, top_col) * 100, 2)

    # Plausibilitaet (verhindert stilles Verrutschen von Zeilen/Spalten).
    if not 7000 <= arbeiter["brutto"] <= 10000:
        fail(f"Median-Bruttoeinkommen unplausibel: {arbeiter['brutto']}")
    if not 8000 <= mittelstand["brutto"] <= 12000:
        fail(f"Durchschnitts-Bruttoeinkommen unplausibel: {mittelstand['brutto']}")
    if not arbeiter["brutto"] < mittelstand["brutto"]:
        fail("Median sollte unter dem Durchschnitt liegen.")
    for h in (arbeiter, mittelstand):
        if not 5 <= h["steuern_pct"] <= 15 or not 1 <= h["vermoegen_pct"] <= 8:
            fail(f"Anteile unplausibel: {h}")

    out = {
        "quelle": "bfs_habe",
        "tabelle": "T20.02.01.00.12",
        "periode": "2015–2017",
        "einheit": "CHF/Monat pro Haushalt (Mittelwert)",
        "arbeiter": arbeiter,
        "mittelstand": mittelstand,
        "oberstes_fuenftel_vermoegen_pct": top_vermoegen_pct,
    }
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
        f.write("\n")

    print(f"  [OK] {os.path.relpath(OUT, ROOT)}")
    print(f"       Arbeiter (Median):      Brutto {arbeiter['brutto']:.0f}, "
          f"Vermögen {arbeiter['vermoegen_pct']} %, Steuern {arbeiter['steuern_pct']} % "
          f"-> {arbeiter['jahre']} Jahre / {arbeiter['tage']:.0f} Tage")
    print(f"       Mittelstand (Schnitt):  Brutto {mittelstand['brutto']:.0f}, "
          f"Vermögen {mittelstand['vermoegen_pct']} %, Steuern {mittelstand['steuern_pct']} % "
          f"-> {mittelstand['jahre']} Jahre / {mittelstand['tage']:.0f} Tage")


if __name__ == "__main__":
    main()
