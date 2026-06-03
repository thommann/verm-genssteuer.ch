import openpyxl, json, math, os
base='/root/.claude/uploads/131eac8a-39f6-47e5-a401-bed7eef8fb5e/'
OUT='/home/user/verm-genssteuer.ch/data/'
def wb(f): return openpyxl.load_workbook(base+f, data_only=True)

rech = wb('1a6139ec-Vermoegenssteuer_Rechner.xlsx')
stat = wb('8184d39d-Vermoegensstatistik_CH_2012_2022.xlsx')
glob = wb('8ce767b0-Globale_Vermoegensverteilung_WID_UBS.xlsx')

# ---- Calculator bins from Engine ----
eng=rech['Engine']
bins=[]
for r in range(2, eng.max_row+1):
    lo=eng.cell(r,1).value; hi=eng.cell(r,2).value; mid=eng.cell(r,3).value
    if lo is None: continue
    bins.append({'lo':lo,'hi':hi,'mid':mid,
        'cnt2020':eng.cell(r,5).value,'cnt2021':eng.cell(r,7).value,'cnt2022':eng.cell(r,9).value})
json.dump(bins, open(OUT+'calculator_bins.json','w'), indent=0)
print("bins:",len(bins))

# yearly params + defaults from Steuerrechner
sr=rech['Steuerrechner']
params={
 'defaults':{'schwelle':sr['C4'].value,'exponent':sr['C5'].value,'cap':sr['C6'].value,
   'ankerVermoegen':sr['C7'].value,'ankerSatz':sr['C8'].value,'mPauschal':sr['C9'].value,'rendite':sr['C10'].value},
 'years':{}
}
for i,yr in enumerate([2020,2021,2022]):
    row=16+i
    params['years'][yr]={'f5_10':sr.cell(row,2).value,'f10':sr.cell(row,3).value,'w10':sr.cell(row,4).value,
      'mean10':sr.cell(row,5).value,'alpha':sr.cell(row,6).value,'Ntail':sr.cell(row,7).value,'xmax':sr.cell(row,8).value}
# published results
params['published_revenue']={2020:sr['B23'].value,2021:sr['C23'].value,2022:sr['D23'].value}
params['published_bands']={}
for r,name in [(27,'5-10 Mio'),(28,'10-100 Mio'),(29,'100 Mio-1 Mrd'),(30,'1-10 Mrd'),(31,'>10 Mrd')]:
    params['published_bands'][name]=[sr.cell(r,2).value,sr.cell(r,3).value,sr.cell(r,4).value]
json.dump(params, open(OUT+'calculator_params.json','w'), indent=2)

# ---- VALIDATE revenue engine ----
def make_tax(schwelle,k,cap,anker,target):
    kp=k+1
    basis=target*anker*kp/(schwelle*((anker/schwelle)**kp-1))
    wcap=schwelle*(cap/basis)**(1.0/k)
    def tax(W):
        if W<=schwelle: return 0.0
        if W<=wcap:
            return basis*schwelle/kp*((W/schwelle)**kp-1)
        below=basis*schwelle/kp*((wcap/schwelle)**kp-1)
        return below+cap*(W-wcap)
    return tax,basis,wcap
d=params['defaults']
tax,basis,wcap=make_tax(d['schwelle'],d['exponent'],d['cap'],d['ankerVermoegen'],d['ankerSatz'])
print("basis=%.10f (xlsx 0.00257231) wcap=%.2f (xlsx 3770402679)"%(basis,wcap))
for yr in [2020,2021,2022]:
    key='cnt%d'%yr
    rev=sum(b[key]*tax(b['mid']) for b in bins)
    print("rev %d = %.4f bn (published %.4f)"%(yr,rev/1e9,params['published_revenue'][yr]))
