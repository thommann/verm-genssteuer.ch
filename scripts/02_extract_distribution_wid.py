import openpyxl, json, math
base='/root/.claude/uploads/131eac8a-39f6-47e5-a401-bed7eef8fb5e/'
OUT='/home/user/verm-genssteuer.ch/data/'
def wb(f): return openpyxl.load_workbook(base+f, data_only=True)
stat = wb('8184d39d-Vermoegensstatistik_CH_2012_2022.xlsx')
glob = wb('8ce767b0-Globale_Vermoegensverteilung_WID_UBS.xlsx')
YEARS=list(range(2012,2023))

def read_pivot(ws, year_cols_start):
    classes=[]
    for r in range(3,14):
        lab=ws.cell(r,1).value
        lo=ws.cell(r,2).value; hi=ws.cell(r,3).value; width=ws.cell(r,4).value
        vals={}
        for j,yr in enumerate(range(year_cols_start[0],year_cols_start[1]+1)):
            vals[yr]=ws.cell(r, 5+j).value
        classes.append({'label':lab or ('> 10\'000\'000' if r==13 else ''),'lo':lo,'hi':hi,'width':width,'values':vals})
    return classes

dist={
 'alle_counts':read_pivot(stat['Pivot_CH'],(2012,2022)),
 'alle_wealth':read_pivot(stat['Pivot_CH_Vermoegen'],(2012,2022)),
 'unb_counts':read_pivot(stat['Pivot_CH_unb'],(2020,2022)),
 'unb_wealth':read_pivot(stat['Pivot_CH_Verm_unb'],(2020,2022)),
}
json.dump(dist, open(OUT+'estv_distribution.json','w'), indent=1)

# Compute key Kennzahlen for "unbeschraenkt" (cleaner: residents) per year 2020-2022
def kennzahlen(counts, wealth, years):
    out={}
    for yr in years:
        cnt=[c['values'][yr] for c in counts]
        wlth=[c['values'][yr] for c in wealth]
        lo=[c['lo'] for c in counts]; hi=[c['hi'] for c in counts]; width=[c['width'] for c in counts]
        N=sum(cnt); W=sum(wlth)
        mean=W/N
        # exact shares for top classes
        # class index: 0 zero,1 0-50k,2 50-100k,3 100-200k,4 200-500k,5 500k-1M,6 1-2M,7 2-3M,8 3-5M,9 5-10M,10 >=10M
        share_ge10=wlth[10]/W
        share_ge5=(wlth[9]+wlth[10])/W
        share_ge1=sum(wlth[6:])/W
        cnt_ge10=cnt[10]; cnt_ge5=cnt[9]+cnt[10]; cnt_ge1=sum(cnt[6:])
        # median via interpolation
        def percentile(p):
            target=p*N; cum=0
            for i in range(len(cnt)):
                if cum+cnt[i]>=target:
                    if width[i] in (0,None): return lo[i]
                    return lo[i]+(target-cum)/cnt[i]*width[i]
                cum+=cnt[i]
            return lo[-1]
        out[yr]={'N':N,'W':W,'mean':mean,'median':percentile(0.5),'p90':percentile(0.9),'p95':percentile(0.95),'p99':percentile(0.99),
          'share_ge10M':share_ge10,'share_ge5M':share_ge5,'share_ge1M':share_ge1,
          'cnt_ge10M':cnt_ge10,'cnt_ge5M':cnt_ge5,'cnt_ge1M':cnt_ge1,
          'pct_ge10M':cnt_ge10/N,'pct_ge5M':cnt_ge5/N,'pct_ge1M':cnt_ge1/N}
    return out
kenn_unb=kennzahlen(dist['unb_counts'],dist['unb_wealth'],[2020,2021,2022])
kenn_alle=kennzahlen(dist['alle_counts'],dist['alle_wealth'],YEARS)
json.dump({'unbeschraenkt':kenn_unb,'alle':kenn_alle}, open(OUT+'estv_kennzahlen.json','w'), indent=1)
print("2022 unb: N=%d W=%.1fbn mean=%.0f median=%.0f share>=10M=%.3f cnt>=10M=%d top1-threshold-ish"%(
  kenn_unb[2022]['N'],kenn_unb[2022]['W']/1e9,kenn_unb[2022]['mean'],kenn_unb[2022]['median'],
  kenn_unb[2022]['share_ge10M'],kenn_unb[2022]['cnt_ge10M']))
print("share>=5M 2022:",round(kenn_unb[2022]['share_ge5M'],4)," cnt>=5M:",kenn_unb[2022]['cnt_ge5M']," pct>=5M:",round(kenn_unb[2022]['pct_ge5M'],4))
print("share>=1M 2022:",round(kenn_unb[2022]['share_ge1M'],4)," pct>=1M:",round(kenn_unb[2022]['pct_ge1M'],4))

# ---- WID time series ----
COUNTRIES=['USA','China','Schweiz','Frankreich','Deutschland','Vereinigtes Königreich','Schweden','Japan','Italien','Spanien','Indien','Russland','Brasilien','Welt']
def read_wid(sheetname):
    ws=glob[sheetname]; series={c:{} for c in COUNTRIES}
    for r in range(5,35):
        yr=ws.cell(r,1).value
        if yr is None: continue
        for j,c in enumerate(COUNTRIES):
            series[c][yr]=ws.cell(r,2+j).value
    return series
wid={'top1':read_wid('WID_Top1_Zeitreihe'),'top10':read_wid('WID_Top10_Zeitreihe'),
     'mid40':read_wid('WID_Mid40_Zeitreihe'),'bot50':read_wid('WID_Bot50_Zeitreihe')}
json.dump(wid, open(OUT+'wid_timeseries.json','w'), indent=1)

# Übersicht latest
ws=glob['Übersicht']; latest=[]
for r in range(6,20):
    if ws.cell(r,1).value is None: continue
    latest.append({'land':ws.cell(r,1).value,'jahr':ws.cell(r,2).value,'top1':ws.cell(r,3).value,
      'top10':ws.cell(r,4).value,'mid40':ws.cell(r,5).value,'bot50':ws.cell(r,6).value,'gini':ws.cell(r,7).value})
json.dump(latest, open(OUT+'wid_latest.json','w'), indent=1)

# UBS gini
ws=glob['UBS_Gini']; ubs=[]
for r in range(5,20):
    if ws.cell(r,1).value is None: continue
    ubs.append({'land':ws.cell(r,1).value,'gini':ws.cell(r,2).value,'hinweis':ws.cell(r,3).value})
json.dump(ubs, open(OUT+'ubs_gini.json','w'), indent=1)

# Pauschal
ws=stat['Pauschalbesteuerung_FDK']
pausch={'counts_ch':{},'revenue':{}}
yrs=[2008,2010,2012,2014,2016,2018]
for j,yr in enumerate(yrs):
    pausch['counts_ch'][yr]=ws.cell(12,2+j).value
for j,yr in enumerate(yrs):
    pausch['revenue'][yr]={'bund':ws.cell(43,2+j).value,'kanton':ws.cell(44,2+j).value,'gemeinde':ws.cell(45,2+j).value}
pausch['lowest']=ws.cell(48,7).value; pausch['highest']=ws.cell(49,7).value
json.dump(pausch, open(OUT+'pauschal.json','w'), indent=1)
print("WID Schweiz top1 2024:",wid['top1']['Schweiz'][2024]," bot50 2024:",wid['bot50']['Schweiz'][2024])
print("done")
