#!/usr/bin/env python3
"""
07: «Wem gehört die Schweiz?» -> src/data/ownership.json

Drei Bausteine fuer die Sektion «Besitz» auf der Verteilungs-Seite:

  1) reichste  Nationalitaet der reichsten Menschen des Landes.
               Kuratierter Wert (kein fetchbarer Cube): «Bilanz – Die 300
               Reichsten» weist je Liste aus, wie viele der 300 Auslaender mit
               Wohnsitz in der Schweiz sind. Liste 2022: 145 von 300 (rund die
               Haelfte), also rund 155 Schweizer bzw. Liechtensteiner. Der Anteil
               ist ueber die Jahrgaenge stabil bei «rund der Haelfte». Quelle und
               Beleg: src/data/sources.json (`bilanz300`), docs/QUELLEN.md §7a.

  2) firmen    Schweizerisch vs. auslaendisch kontrollierte Unternehmensgruppen.
               BFS Statistik der Unternehmensgruppen (STAGRE), Tabelle T 6.6.3
               «Anzahl Unternehmensgruppen, Unternehmen, Beschaeftigte und Umsatz
               nach Sitzland». Vier Blaetter (Gruppen, Unternehmen, Besch., Umsatz);
               Zeile «Schweiz» = inlaendisch kontrolliert, Auslandskontrolle =
               «Total» minus «Schweiz». Je Blatt das neueste Jahr mit beiden Werten.

Quellen (live, ein Befehl ueber scripts/fetch_sources.sh):
  data/raw/bfs/stagre-sitzland.xlsx   (firmen)
Der reichste-Block ist ein belegter, kuratierter Publikationswert (bezug=kuratiert),
analog zu den EFV/BAG/LITRA-Konstanten in 04_extract_spend_reference.py.

Verwendung:
  bash scripts/fetch_sources.sh
  python3 scripts/07_extract_ownership.py

Benoetigt openpyxl.
"""
import csv
import json
import os
import re
import sys

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW = os.path.join(ROOT, "data", "raw")
STAGRE = os.path.join(RAW, "bfs", "stagre-sitzland.xlsx")
GEBAEUDE = os.path.join(RAW, "bfs", "gebaeude-eigentuemertyp.xlsx")
WOHNEIGENTUM = os.path.join(RAW, "bfs", "wohneigentumsquote.xlsx")
SNB = os.path.join(RAW, "snb", "haushalte-vermoegen.csv")
GDP = os.path.join(RAW, "worldbank", "che_gdp_current_chf.json")
OUT = os.path.join(ROOT, "src", "data", "ownership.json")

# Blatt -> Schluessel in der Ausgabe (Reihenfolge wie im UI).
STAGRE_SHEETS = [
    ("Unternehmen", "unternehmen"),
    ("Besch.", "beschaeftigte"),
    ("Umsatz", "umsatz"),
    ("Gruppen", "gruppen"),
]

# Kuratierter Beleg «Bilanz – Die 300 Reichsten», Liste 2022 (belegt in
# sources.json `bilanz300`, docs/QUELLEN.md §7a): 145 der 300 Reichsten sind
# Auslaender mit Wohnsitz in der Schweiz, also rund 155 Schweizer/Liechtensteiner.
BILANZ_REICHSTE = {"jahr": 2022, "total": 300, "auslaender": 145}
# Gesamtvermoegen der 300 Reichsten, Liste 2025 (Bilanz, kuratiert; vgl. bilanz300).
BILANZ_VERMOEGEN_MRD = 851.5
BILANZ_VERMOEGEN_JAHR = 2025
# Gesamtvermoegen der 300 Reichsten je Liste, in Mrd. CHF (Bilanz, kuratiert; frei belegt):
# 2022 = 821 (telebasel/SRF), 2023 = 795 (watson), 2024 = 833,5 (SRF u.a.), 2025 = 851,5.
BILANZ_SERIE = {2022: 821.0, 2023: 795.0, 2024: 833.5, 2025: 851.5}
# Kuratierter Beleg Raiffeisen Economic Research «Immobilien Schweiz Q4/20»
# (Quelle `raiffeisen_immo`): Eigentuemeranteile am Mietwohnungsbestand der Schweiz.
MIETWOHNUNGEN = {
    "jahr": 2020,
    "privat": 0.49, "institutionell": 0.33,
    "genossenschaften": 0.08, "immobilienfirmen": 0.07, "oeffentlich": 0.04,
}
# Indirekter Auslandsbesitz trotz Lex Koller: BlackRock an boersenkotierten Schweizer
# Immobilienfirmen (REFLEKT/WAV), kuratiert. Quelle `blackrock_immo`.
BLACKROCK = {"anteil": 0.06, "wert_mrd": 2.0, "firmen": 17, "faktor_10j": 20, "sps_anteil": 0.12}
# BFS-Wohnungsstatistik (GWS, kumuliert 2021-2023), kuratiert. Quelle `bfs_wohnungen`.
# Privatanteil an den Mietwohnungen und nach Bauperiode.
WOHNUNGEN_BFS = {
    "jahr": 2023, "privat": 0.45,
    "baujahr": {"vor_1946": 0.65, "nach_2000": 0.32},
}
# Pachtanteil an der landwirtschaftlichen Nutzflaeche (BFS-Strukturerhebung), kuratiert.
# Quelle `bfs_pacht`. Steigender Trend mit dem Strukturwandel.
PACHT_SERIE = [
    {"jahr": 1980, "anteil": 0.37}, {"jahr": 2005, "anteil": 0.43},
    {"jahr": 2016, "anteil": 0.45}, {"jahr": 2020, "anteil": 0.47},
]
# Groesste einzelne Grundbesitzer nach Flaeche bzw. Wohnungsbestand (Medienrecherchen),
# kuratiert. menge/einheit + Quelle je Beispiel.
GROESSTE = [
    {"id": "ubs", "name": "UBS (mit CS)", "menge": 70000, "einheit": "Wohnungen", "quelle": "ubs_wohnungen"},
    {"id": "swisslife", "name": "Swiss Life", "menge": 3, "einheit": "Mio. m²", "quelle": "bilanz_immo"},
    {"id": "hiag", "name": "Hiag (Familie Grisard)", "menge": 2.6, "einheit": "km²", "quelle": "hiag"},
]


def fail(msg):
    print(f"  [FAIL] {msg}", file=sys.stderr)
    sys.exit(1)


# ---------------------------------------------------------------------------
# 1) Reichste: Nationalitaet der 300 Reichsten (kuratiert, Bilanz) plus ihr
#    Anteil am gesamten privaten Reinvermoegen der Schweiz (SNB, fetchbar).
# ---------------------------------------------------------------------------
def parse_snb_reinvermoegen():
    """Jaehrliches Reinvermoegen der privaten Haushalte (SNB-Cube frsekgevehup,
    Code RVM, in Mio. CHF) -> {jahr: mrd}."""
    if not os.path.exists(SNB):
        fail(f"SNB-Datei fehlt: {SNB}; zuerst `bash scripts/fetch_sources.sh`.")
    out = {}
    with open(SNB, encoding="utf-8-sig") as fh:
        for row in csv.reader(fh, delimiter=";"):
            if len(row) != 3 or row[1] != "RVM":
                continue
            try:
                out[int(row[0])] = float(row[2]) / 1000.0  # Mio. -> Mrd.
            except ValueError:
                continue
    if not out:
        fail("SNB: keine RVM-Zeile (Reinvermoegen) gefunden.")
    return out


def parse_gdp():
    """Nominales BIP der Schweiz je Jahr (Weltbank NY.GDP.MKTP.CN, CHF) -> {jahr: mrd}."""
    if not os.path.exists(GDP):
        fail(f"BIP-Datei fehlt: {GDP}; zuerst `bash scripts/fetch_sources.sh`.")
    with open(GDP, encoding="utf-8") as fh:
        d = json.load(fh)
    out = {}
    for r in d[1]:
        if r.get("value") is not None:
            out[int(r["date"])] = r["value"] / 1e9  # CHF -> Mrd.
    if not out or max(out.values()) < 400:
        fail("BIP-Reihe (Weltbank) nicht plausibel.")
    return out


def build_reichste():
    total = BILANZ_REICHSTE["total"]
    auslaender = BILANZ_REICHSTE["auslaender"]
    schweizer = total - auslaender
    if not (0 < auslaender < total):
        fail(f"Reichste: Auslaenderzahl unplausibel ({auslaender}/{total}).")
    rvm = parse_snb_reinvermoegen()
    gdp = parse_gdp()
    snb_jahr = max(rvm)
    privat_mrd = rvm[snb_jahr]
    anteil = BILANZ_VERMOEGEN_MRD / privat_mrd
    if not (3000 <= privat_mrd <= 8000):
        fail(f"SNB-Reinvermoegen unplausibel: {privat_mrd:.0f} Mrd.")
    if not (0.1 <= anteil <= 0.3):
        fail(f"Vermoegensanteil der 300 unplausibel: {anteil:.3f}.")

    # Trend A: gesamtes Privatvermoegen als Vielfaches des BIP (SNB/Weltbank), je Jahr.
    vermoegen_bip = [
        {"jahr": y, "vielfaches": round(rvm[y] / gdp[y], 2)}
        for y in sorted(rvm) if y in gdp
    ]
    # Trend B: die 300 Reichsten als Anteil am BIP (Bilanz/Weltbank), je Listenjahr.
    top300_bip = [
        {"jahr": y, "anteil": round(BILANZ_SERIE[y] / gdp[y], 4)}
        for y in sorted(BILANZ_SERIE) if y in gdp
    ]
    if len(vermoegen_bip) < 10 or vermoegen_bip[-1]["vielfaches"] < 4:
        fail("Vermoegen/BIP-Reihe nicht plausibel.")
    if len(top300_bip) < 2:
        fail("300/BIP-Reihe zu kurz (BIP-Jahre fehlen).")
    return {
        "quelle": "bilanz300",
        "bezug": "kuratiert",
        "jahr": BILANZ_REICHSTE["jahr"],
        "total": total,
        "auslaender": auslaender,
        "schweizer": schweizer,
        "auslaender_share": round(auslaender / total, 4),
        "schweizer_share": round(schweizer / total, 4),
        # Vermoegenskonzentration: 300 Reichste (Bilanz, kuratiert) am gesamten
        # privaten Reinvermoegen der Schweiz (SNB, fetchbar).
        "vermoegen_300_mrd": BILANZ_VERMOEGEN_MRD,
        "vermoegen_300_jahr": BILANZ_VERMOEGEN_JAHR,
        "privatvermoegen_mrd": round(privat_mrd, 1),
        "privatvermoegen_jahr": snb_jahr,
        "anteil_privatvermoegen": round(anteil, 4),
        # Verlaeufe gegenueber dem BIP.
        "vermoegen_bip_serie": vermoegen_bip,
        "top300_bip_serie": top300_bip,
    }


# ---------------------------------------------------------------------------
# 2) Firmen: STAGRE T 6.6.3, je Blatt Total und Schweiz, neuestes Jahr.
# ---------------------------------------------------------------------------
def parse_firmen():
    if not os.path.exists(STAGRE):
        fail(f"STAGRE-Datei fehlt: {STAGRE}; zuerst `bash scripts/fetch_sources.sh`.")
    wb = openpyxl.load_workbook(STAGRE, data_only=True)
    out = {"quelle": "bfs_stagre"}

    def row_by_label(ws, label):
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if isinstance(v, str) and v.strip() == label:
                return [ws.cell(r, c).value for c in range(2, ws.max_column + 1)]
        return None

    for sheet, key in STAGRE_SHEETS:
        if sheet not in wb.sheetnames:
            fail(f"STAGRE-Blatt «{sheet}» fehlt; vorhanden: {wb.sheetnames}")
        ws = wb[sheet]
        years = row_by_label(ws, "Sitzland")  # Header-Zeile traegt die Jahre.
        total = row_by_label(ws, "Total")
        ch = row_by_label(ws, "Schweiz")
        if not years or not total or not ch:
            fail(f"STAGRE «{sheet}»: Header/Total/Schweiz-Zeile nicht gefunden.")
        # Neuestes Jahr (von rechts), bei dem Total UND Schweiz numerisch sind.
        idx = None
        for i in range(len(total) - 1, -1, -1):
            t, c = total[i], ch[i]
            if isinstance(t, (int, float)) and isinstance(c, (int, float)) \
                    and isinstance(years[i], (int, float)):
                idx = i
                break
        if idx is None:
            fail(f"STAGRE «{sheet}»: kein Jahr mit Total und Schweiz.")
        t, c = float(total[idx]), float(ch[idx])
        foreign = t - c
        if not (t > 0 and 0 < c < t):
            fail(f"STAGRE «{sheet}»: Werte unplausibel (Total={t}, Schweiz={c}).")
        # Zeitreihe des Auslandskontroll-Anteils ueber alle Jahre mit Werten.
        serie = []
        for i in range(len(total)):
            ti, ci, yi = total[i], ch[i], years[i]
            if isinstance(ti, (int, float)) and isinstance(ci, (int, float)) \
                    and isinstance(yi, (int, float)) and ti > 0:
                serie.append({"jahr": int(yi), "ausland_share": round((ti - ci) / ti, 4)})
        out[key] = {
            "jahr": int(years[idx]),
            "total": round(t, 1),
            "ch": round(c, 1),
            "ausland": round(foreign, 1),
            "ch_share": round(c / t, 4),
            "ausland_share": round(foreign / t, 4),
            "serie": serie,
        }
    # Selbstpruefung: auslaendische Gruppen tragen einen kleinen Firmen-/Job-
    # Anteil, aber den groesseren Umsatzanteil (Kernaussage der Sektion).
    if out["umsatz"]["ausland_share"] <= out["unternehmen"]["ausland_share"]:
        fail("STAGRE: Umsatzanteil Ausland sollte ueber dem Firmenanteil liegen.")
    return out


def build_boden_kuratiert():
    """Kuratierte Boden-Zusatzwerte: indirekter Auslandsbesitz BlackRock (REFLEKT)
    und groesste einzelne Eigentuemer (Medienrecherchen)."""
    if not GROESSTE or any(g["menge"] <= 0 for g in GROESSTE):
        fail("Groesste-Eigentuemer-Liste unplausibel.")
    return {
        "blackrock": {"quelle": "blackrock_immo", "bezug": "kuratiert", **BLACKROCK},
        "pacht": {"quelle": "bfs_pacht", "bezug": "kuratiert", "serie": PACHT_SERIE},
        "groesste": [{"bezug": "kuratiert", **g} for g in GROESSTE],
    }


def parse_wohneigentum():
    """Wohneigentumsquote der Schweiz je Jahr (BFS-Tabelle T 09.03.02.01.03, ein Blatt je
    Jahr, Zeile «Schweiz», Spalte «Wohneigentumsquote / Anteil in %»)."""
    if not os.path.exists(WOHNEIGENTUM):
        fail(f"Wohneigentums-Datei fehlt: {WOHNEIGENTUM}; zuerst `bash scripts/fetch_sources.sh`.")
    wb = openpyxl.load_workbook(WOHNEIGENTUM, data_only=True)
    serie = []
    for sn in wb.sheetnames:
        if not sn.isdigit():
            continue
        ws = wb[sn]
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value and str(ws.cell(r, 1).value).strip() == "Schweiz":
                q = ws.cell(r, 14).value  # Spalte «Anteil in %»
                if isinstance(q, (int, float)) and 20 <= q <= 50:  # plausible Quote
                    serie.append({"jahr": int(sn), "quote": round(q / 100, 4)})
                break
    serie.sort(key=lambda p: p["jahr"])
    if len(serie) < 8:
        fail(f"Wohneigentumsquote: zu wenige Jahre ({len(serie)}).")
    peak = max(serie, key=lambda p: p["quote"])
    return {
        "quelle": "bfs_wohneigentum",
        "serie": serie,
        "jahr": serie[-1]["jahr"], "quote": serie[-1]["quote"],
        "peak_jahr": peak["jahr"], "peak_quote": peak["quote"],
    }


def build_wohnungen_bfs():
    """Mietwohnungseigentum BFS (kuratiert): Privatanteil und Bauperioden-Trend."""
    w = WOHNUNGEN_BFS
    if not (0.3 <= w["privat"] <= 0.6 and w["baujahr"]["vor_1946"] > w["baujahr"]["nach_2000"]):
        fail("BFS-Wohnungen: Werte unplausibel.")
    return {"quelle": "bfs_wohnungen", "bezug": "kuratiert", **w}


# ---------------------------------------------------------------------------
# 4) Gebaeude nach Eigentuemertyp (BFS, registerbasiert GWR + Grundbuch).
# ---------------------------------------------------------------------------
def parse_gebaeude():
    if not os.path.exists(GEBAEUDE):
        fail(f"Gebaeude-Datei fehlt: {GEBAEUDE}; zuerst `bash scripts/fetch_sources.sh`.")
    wb = openpyxl.load_workbook(GEBAEUDE, data_only=True)
    sheet = "2022" if "2022" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet]
    # Zeile «Total (Schweiz ohne ZH und VS…)». Spalten (Header rows 3-6):
    # 2=Total, 3=Natuerliche Person(en), 4=Juristische Person (Total),
    # 15=Gemeinschaft (Total), 20=Gemischt, 21=Unbekannt.
    row = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip().startswith("Total (Schweiz"):
            row = r
            break
    if row is None:
        fail("Gebaeude: «Total (Schweiz …)»-Zeile nicht gefunden.")

    def cell(c):
        v = ws.cell(row, c).value
        if not isinstance(v, (int, float)):
            fail(f"Gebaeude: Zelle Spalte {c} keine Zahl: {v!r}")
        return float(v)

    total = cell(2)
    natuerliche = cell(3)
    juristische = cell(4)
    gemeinschaft = cell(15)
    gemischt = cell(20)
    unbekannt = cell(21)
    summe = natuerliche + juristische + gemeinschaft + gemischt + unbekannt
    if abs(summe - total) / total > 0.005:
        fail(f"Gebaeude: Summe der Eigentuemertypen ({summe}) != Total ({total}).")
    if not (0.6 <= natuerliche / total <= 0.75):
        fail(f"Gebaeude: Anteil natuerliche Personen unplausibel ({natuerliche/total:.3f}).")
    return {
        "quelle": "bfs_gebaeude",
        "jahr": int(sheet) if sheet.isdigit() else 2022,
        "total": round(total),
        "natuerliche": round(natuerliche),
        "juristische": round(juristische),
        "gemeinschaft": round(gemeinschaft),
        "gemischt": round(gemischt),
        "unbekannt": round(unbekannt),
        "natuerliche_share": round(natuerliche / total, 4),
        "juristische_share": round(juristische / total, 4),
        "gemeinschaft_share": round(gemeinschaft / total, 4),
        "uebrige_share": round((gemischt + unbekannt) / total, 4),
    }


def build_mietwohnungen():
    """Eigentuemeranteile am Mietwohnungsbestand (kuratiert, Raiffeisen Q4/20)."""
    m = MIETWOHNUNGEN
    summe = m["privat"] + m["institutionell"] + m["genossenschaften"] \
        + m["immobilienfirmen"] + m["oeffentlich"]
    if abs(summe - 1.0) > 0.02:
        fail(f"Mietwohnungen: Anteile summieren nicht auf 100 % ({summe}).")
    return {"quelle": "raiffeisen_immo", "bezug": "kuratiert", **m}


def main():
    data = {
        "reichste": build_reichste(),
        "firmen": parse_firmen(),
        "gebaeude": parse_gebaeude(),
        "mietwohnungen": build_mietwohnungen(),
        "wohnungen_bfs": build_wohnungen_bfs(),
        "wohneigentum": parse_wohneigentum(),
        "boden": build_boden_kuratiert(),
    }
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")

    re_, fi, ge = data["reichste"], data["firmen"], data["gebaeude"]
    print(f"  [OK] {os.path.relpath(OUT, ROOT)}")
    print(f"       Reichste ({re_['jahr']}): {re_['auslaender']} von {re_['total']} sind "
          f"Auslaender ({re_['auslaender_share']*100:.1f} %), {re_['schweizer']} Schweizer/"
          f"Liechtensteiner ({re_['schweizer_share']*100:.1f} %)")
    print(f"       Spitze: 300 Reichste {re_['vermoegen_300_mrd']:.0f} Mrd = "
          f"{re_['anteil_privatvermoegen']*100:.1f} % des privaten Reinvermoegens "
          f"({re_['privatvermoegen_mrd']:.0f} Mrd, SNB {re_['privatvermoegen_jahr']})")
    print(f"       Firmen ({fi['unternehmen']['jahr']}): Ausland {fi['unternehmen']['ausland_share']*100:.1f} % "
          f"der Unternehmen, {fi['beschaeftigte']['ausland_share']*100:.1f} % der Jobs, "
          f"{fi['umsatz']['ausland_share']*100:.1f} % des Umsatzes ({fi['umsatz']['jahr']})")
    print(f"       Gebaeude ({ge['jahr']}): {ge['natuerliche_share']*100:.1f} % natuerliche Personen, "
          f"{ge['juristische_share']*100:.1f} % juristische, {ge['gemeinschaft_share']*100:.1f} % Gemeinschaften")
    mw = data["mietwohnungen"]
    print(f"       Mietwohnungen ({mw['jahr']}): privat {mw['privat']*100:.0f} %, "
          f"institutionell {mw['institutionell']*100:.0f} %, Genossenschaften {mw['genossenschaften']*100:.0f} %")


if __name__ == "__main__":
    main()
